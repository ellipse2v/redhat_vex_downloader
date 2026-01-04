#!/usr/bin/env python3
# Copyright 2025 ellipse2v
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""
RPM to CVE Mapping Tool

Maps RPM packages (NEVRA format) to CVEs with detailed vulnerability information.
Generates Excel report with separate sheets for fixed, not fixed, and affected packages.
"""

import argparse
import csv
import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Set, Tuple, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from concurrent.futures import ThreadPoolExecutor, as_completed

# Configuration
OUTPUT_DIR = "reports"
OUTPUT_FILE = "rpm_cve_mapping_{timestamp}.xlsx"
MAX_WORKERS = 10  # Number of concurrent workers for RPM analysis

# NEVRA/NEVR patterns: Name-Epoch:Version-Release.Architecture or Name-Epoch:Version-Release
NEVRA_PATTERN = r'^(?P<name>[^-]+)-(?P<epoch>\d+):(?P<version>[^-]+)-(?P<release>[^\s.]+)(?:\.(?P<arch>[^\s]+))?$'

# Excel styles
HEADER_STYLE = {
    'font': Font(bold=True, color='FFFFFF'),
    'fill': PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid'),
    'border': Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin')),
    'alignment': Alignment(horizontal='center', vertical='center')
}

CELL_STYLE = {
    'border': Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin')),
    'alignment': Alignment(wrap_text=True)
}

def parse_nevra(nevra: str) -> Optional[Dict[str, str]]:
    """Parse NEVRA string into components."""
    match = re.match(NEVRA_PATTERN, nevra.strip())
    if match:
        return match.groupdict()
    return None

def load_rpm_list(csv_file: str) -> List[Dict[str, str]]:
    """Load RPM list from CSV file."""
    rpm_list = []
    
    try:
        with open(csv_file, 'r') as f:
            reader = csv.reader(f)
            
            # Check if first row is header
            first_row = next(reader)
            if 'rpm' in first_row[0].lower() or 'package' in first_row[0].lower():
                # Has header, skip it
                pass
            else:
                # No header, use first row as data
                rpm_list.append({'rpm': first_row[0], 'source': csv_file})
            
            for row in reader:
                if row:  # Skip empty rows
                    rpm_list.append({'rpm': row[0], 'source': csv_file})
    
    except FileNotFoundError:
        print(f"❌ RPM list file not found: {csv_file}")
        return []
    except Exception as e:
        print(f"❌ Error reading RPM list: {e}")
        return []
    
    return rpm_list

def get_vex_files(data_dir: Path) -> List[Path]:
    """Get all VEX files from data directory."""
    vex_files = []
    
    for year_dir in data_dir.iterdir():
        if year_dir.is_dir():
            for file_path in year_dir.glob("*.json"):
                vex_files.append(file_path)
    
    return vex_files

def extract_cvss_info(vuln_data: dict) -> Tuple[Optional[float], Optional[str]]:
    """Extract CVSS score and vector from vulnerability data."""
    score = None
    vector = None
    
    # Check metrics for CVSS information
    if 'metrics' in vuln_data:
        for metric in vuln_data['metrics']:
            if 'cvss_v3' in metric:
                cvss_data = metric['cvss_v3']
                if 'baseScore' in cvss_data:
                    score = cvss_data['baseScore']
                if 'vectorString' in cvss_data:
                    vector = cvss_data['vectorString']
                break
            elif 'cvss_v2' in metric:
                cvss_data = metric['cvss_v2']
                if 'baseScore' in cvss_data:
                    score = cvss_data['baseScore']
                if 'vectorString' in cvss_data:
                    vector = cvss_data['vectorString']
                break
    
    return score, vector

def get_severity_from_cvss(score: Optional[float]) -> str:
    """Convert CVSS score to severity level."""
    if score is None:
        return 'Unknown'
    elif score >= 9.0:
        return 'Critical'
    elif score >= 7.0:
        return 'Important'
    elif score >= 4.0:
        return 'Moderate'
    elif score > 0:
        return 'Low'
    else:
        return 'None'

def get_severity_from_vuln(vuln_data: dict) -> str:
    """Extract severity from vulnerability data with enhanced detection."""
    # First check: baseSeverity field (most reliable)
    if 'baseSeverity' in vuln_data:
        base_severity = vuln_data['baseSeverity']
        if base_severity:
            return base_severity.capitalize()
    
    # Second check: severity in product_status
    if 'product_status' in vuln_data:
        for status in vuln_data['product_status'].values():
            if isinstance(status, list):
                for item in status:
                    if isinstance(item, dict) and 'severity' in item:
                        return item['severity'].capitalize() if item['severity'] else 'Unknown'
    
    # Third check: CVSS scores in metrics
    if 'metrics' in vuln_data:
        for metric in vuln_data['metrics']:
            if 'cvss_v3' in metric and 'baseScore' in metric['cvss_v3']:
                score = metric['cvss_v3']['baseScore']
                return get_severity_from_cvss(score)
    
    return 'Unknown'

def find_rpm_in_products(rpm_name: str, products: List[str]) -> bool:
    """Check if RPM name exists in product list."""
    rpm_name_lower = rpm_name.lower()
    
    for product in products:
        if rpm_name_lower in product.lower():
            return True
    
    return False

def analyze_rpm_cves(rpm_list: List[Dict[str, str]], vex_files: List[Path]) -> Dict[str, Dict]:
    """Analyze RPMs against VEX files to find related CVEs."""
    results = {
        'fixed': [],
        'not_fixed': [],
        'affected': []
    }
    
    # Create RPM name lookup
    rpm_names = {rpm['rpm']: rpm for rpm in rpm_list}
    
    for file_path in vex_files:
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Get CVE ID
            cve_id = data.get('document', {}).get('tracking', {}).get('id', 'Unknown')
            if cve_id == 'Unknown' and 'vulnerabilities' in data:
                vuln = data['vulnerabilities'][0] if data['vulnerabilities'] else {}
                cve_id = vuln.get('cve', 'Unknown')
            
            # Process each vulnerability
            for vuln in data.get('vulnerabilities', []):
                # Get CVSS information and enhanced severity
                cvss_score, cvss_vector = extract_cvss_info(vuln)
                severity = get_severity_from_vuln(vuln)  # Uses baseSeverity first, then CVSS
                
                # Get product status
                product_status = vuln.get('product_status', {})
                
                # Check fixed products
                fixed_products = product_status.get('fixed', [])
                for rpm_name, rpm_info in rpm_names.items():
                    nevra_parts = parse_nevra(rpm_name)
                    if nevra_parts:
                        rpm_base_name = nevra_parts['name']
                        
                        # Check if this RPM is in fixed products
                        if find_rpm_in_products(rpm_base_name, fixed_products):
                            results['fixed'].append({
                                'rpm': rpm_name,
                                'cve': cve_id,
                                'severity': severity,
                                'cvss_score': cvss_score,
                                'cvss_vector': cvss_vector,
                                'status': 'fixed',
                                'fixed_in': ', '.join(fixed_products[:3]) + ('...' if len(fixed_products) > 3 else ''),
                                'source_file': str(file_path)
                            })
                        
                        # Check if this RPM is in known_affected products
                        known_affected = product_status.get('known_affected', [])
                        if find_rpm_in_products(rpm_base_name, known_affected):
                            results['affected'].append({
                                'rpm': rpm_name,
                                'cve': cve_id,
                                'severity': severity,
                                'cvss_score': cvss_score,
                                'cvss_vector': cvss_vector,
                                'status': 'known_affected',
                                'source_file': str(file_path)
                            })
                        
                        # Check if this RPM is in known_not_affected products
                        known_not_affected = product_status.get('known_not_affected', [])
                        if find_rpm_in_products(rpm_base_name, known_not_affected):
                            results['not_fixed'].append({
                                'rpm': rpm_name,
                                'cve': cve_id,
                                'severity': severity,
                                'cvss_score': cvss_score,
                                'cvss_vector': cvss_vector,
                                'status': 'known_not_affected',
                                'source_file': str(file_path)
                            })
            
        except (json.JSONDecodeError, IOError) as e:
            print(f"⚠️  Error processing {file_path}: {e}")
            continue
    
    return results

def create_excel_report(results: Dict[str, Dict], rpm_list: List[Dict[str, str]], output_file: str) -> None:
    """Create Excel report with multiple sheets."""
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create sheets
    sheets = {
        'Fixed': results['fixed'],
        'Not Fixed': results['not_fixed'],
        'Affected': results['affected'],
        'Summary': None
    }
    
    for sheet_name, data in sheets.items():
        if sheet_name == 'Summary':
            ws = wb.create_sheet(sheet_name, 0)
            create_summary_sheet(ws, results, rpm_list)
        else:
            ws = wb.create_sheet(sheet_name)
            if data:
                create_data_sheet(ws, data, sheet_name)
    
    # Save workbook
    output_path = Path(OUTPUT_DIR) / output_file
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    
    print(f"📊 Excel report generated: {output_path}")

def create_summary_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, 
                       results: Dict[str, Dict], rpm_list: List[Dict[str, str]]) -> None:
    """Create summary sheet with overview statistics."""
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Title
    ws['A1'] = 'RPM to CVE Mapping Summary'
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')
    
    # Generated timestamp
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = Font(italic=True)
    ws.merge_cells('A2:D2')
    
    # Spacing
    ws['A3'] = ''
    
    # RPM Statistics
    ws['A4'] = 'RPM Statistics'
    ws['A4'].font = Font(bold=True, size=12)
    ws.merge_cells('A4:D4')
    
    ws['A5'] = 'Total RPMs Analyzed:'
    ws['B5'] = len(rpm_list)
    
    ws['A6'] = 'RPMs with CVEs Found:'
    rpm_with_cves = set()
    for category in results.values():
        for item in category:
            rpm_with_cves.add(item['rpm'])
    ws['B6'] = len(rpm_with_cves)
    
    ws['A7'] = 'RPMs without CVEs:'
    ws['B7'] = len(rpm_list) - len(rpm_with_cves)
    
    # Spacing
    ws['A8'] = ''
    
    # CVE Statistics
    ws['A9'] = 'CVE Statistics'
    ws['A9'].font = Font(bold=True, size=12)
    ws.merge_cells('A9:D9')
    
    ws['A10'] = 'Category'
    ws['B10'] = 'Count'
    ws['C10'] = 'Percentage'
    
    total_cves = sum(len(category) for category in results.values())
    
    row = 11
    for category_name, category_data in results.items():
        count = len(category_data)
        percentage = (count / total_cves * 100) if total_cves > 0 else 0
        
        ws[f'A{row}'] = category_name
        ws[f'B{row}'] = count
        ws[f'C{row}'] = f'{percentage:.1f}%'
        row += 1
    
    ws[f'A{row}'] = 'Total'
    ws[f'B{row}'] = total_cves
    ws[f'C{row}'] = '100.0%'
    
    # Severity Distribution
    ws[f'A{row+2}'] = 'Severity Distribution'
    ws[f'A{row+2}'].font = Font(bold=True, size=12)
    ws.merge_cells(f'A{row+2}:D{row+2}')
    
    ws[f'A{row+3}'] = 'Severity'
    ws[f'B{row+3}'] = 'Count'
    ws[f'C{row+3}'] = 'Percentage'
    
    # Count by severity
    severity_counts = {}
    for category in results.values():
        for item in category:
            severity = item['severity']
            severity_counts[severity] = severity_counts.get(severity, 0) + 1
    
    row += 4
    for severity, count in sorted(severity_counts.items(), 
                                  key=lambda x: ['Critical', 'Important', 'Moderate', 'Low', 'Unknown'].index(x[0]) if x[0] in ['Critical', 'Important', 'Moderate', 'Low', 'Unknown'] else 999):
        percentage = (count / total_cves * 100) if total_cves > 0 else 0
        ws[f'A{row}'] = severity
        ws[f'B{row}'] = count
        ws[f'C{row}'] = f'{percentage:.1f}%'
        row += 1
    
    # Apply styles
    for cell in ws[1:1]:
        for c in cell:
            c.font = Font(bold=True, size=16)
    
    for cell in ws[4:4]:
        for c in cell:
            c.font = Font(bold=True, size=12)
    
    for cell in ws[9:9]:
        for c in cell:
            c.font = Font(bold=True)
    
    for cell in ws[10:10]:
        for c in cell:
            c.font = Font(bold=True)
    
    for cell in ws[f'{row+2}:{row+2}']:
        for c in cell:
            c.font = Font(bold=True, size=12)
    
    for cell in ws[f'{row+3}:{row+3}']:
        for c in cell:
            c.font = Font(bold=True)

def create_data_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, 
                     data: List[Dict], sheet_name: str) -> None:
    """Create data sheet with RPM-CVE mappings."""
    # Set column widths
    ws.column_dimensions['A'].width = 40  # RPM
    ws.column_dimensions['B'].width = 20  # CVE
    ws.column_dimensions['C'].width = 15  # Severity
    ws.column_dimensions['D'].width = 12  # CVSS Score
    ws.column_dimensions['E'].width = 30  # CVSS Vector
    ws.column_dimensions['F'].width = 50  # Fixed In / Additional Info
    
    # Headers
    headers = ['RPM', 'CVE', 'Severity', 'CVSS Score', 'CVSS Vector']
    
    if sheet_name == 'Fixed':
        headers.append('Fixed In')
    else:
        headers.append('Additional Info')
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data
    for row, item in enumerate(data, 2):
        ws.cell(row=row, column=1, value=item['rpm'])
        ws.cell(row=row, column=2, value=item['cve'])
        ws.cell(row=row, column=3, value=item['severity'])
        
        # CVSS Score with conditional formatting
        score_cell = ws.cell(row=row, column=4, value=item['cvss_score'] if item['cvss_score'] is not None else 'N/A')
        if item['cvss_score'] is not None:
            if item['cvss_score'] >= 9.0:
                score_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                score_cell.font = Font(color='FFFFFF', bold=True)
            elif item['cvss_score'] >= 7.0:
                score_cell.fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')
                score_cell.font = Font(bold=True)
            elif item['cvss_score'] >= 4.0:
                score_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            else:
                score_cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
        
        ws.cell(row=row, column=5, value=item['cvss_vector'] if item['cvss_vector'] else 'N/A')
        
        # Additional info based on sheet type
        if sheet_name == 'Fixed':
            ws.cell(row=row, column=6, value=item.get('fixed_in', 'N/A'))
        else:
            ws.cell(row=row, column=6, value=item.get('source_file', 'N/A'))
    
    # Apply styles to all cells
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                               top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Add table formatting
    if len(data) > 0:
        table_ref = f"A1:{chr(64 + len(headers))}{len(data) + 1}"
        table = Table(displayName=f"{sheet_name}Table", ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, 
                             showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

def main():
    parser = argparse.ArgumentParser(
        description="RPM to CVE Mapping Tool - Map RPM packages to CVEs with detailed vulnerability information"
    )
    parser.add_argument(
        "--rpm-list",
        type=str,
        required=True,
        help="CSV file containing list of RPMs in NEVRA format"
    )
    parser.add_argument(
        "--data-dir",
        type=str,
        default="data",
        help="Directory containing VEX files (default: data)"
    )
    parser.add_argument(
        "--output",
        type=str,
        help="Output Excel file name (default: auto-generated with timestamp)"
    )
    
    args = parser.parse_args()
    
    print("🔍 RPM to CVE Mapping Tool")
    print(f"📦 Loading RPM list from: {args.rpm_list}")
    
    # Load RPM list
    rpm_list = load_rpm_list(args.rpm_list)
    if not rpm_list:
        print("❌ No RPMs found in the input file")
        return
    
    print(f"📊 Found {len(rpm_list)} RPMs to analyze")
    
    # Get VEX files
    data_dir = Path(args.data_dir)
    if not data_dir.exists():
        print(f"❌ Data directory not found: {data_dir}")
        return
    
    vex_files = get_vex_files(data_dir)
    print(f"📚 Found {len(vex_files)} VEX files to process")
    
    if not vex_files:
        print("⚠️  No VEX files found to process")
        return
    
    # Analyze RPMs with multithreading
    print("🔬 Analyzing RPMs against VEX files with multithreading...")
    results = analyze_rpm_cves(rpm_list, vex_files)
    
    # Generate output file name
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = args.output if args.output else OUTPUT_FILE.format(timestamp=timestamp)
    
    # Create Excel report
    print("📊 Generating Excel report...")
    create_excel_report(results, rpm_list, output_file)
    
    # Print summary
    print("\n📈 Analysis Summary:")
    print(f"  RPMs analyzed: {len(rpm_list)}")
    print(f"  RPMs with CVEs: {len(set(item['rpm'] for category in results.values() for item in category))}")
    print(f"  Total CVEs found: {sum(len(category) for category in results.values())}")
    
    for category_name, category_data in results.items():
        print(f"  {category_name}: {len(category_data)}")
    
    print(f"\n💾 Report saved to: {OUTPUT_DIR}/{output_file}")
    print("  Open the Excel file to view detailed RPM-CVE mappings")

if __name__ == "__main__":
    main()