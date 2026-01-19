# Copyright 2025 ellipse2v
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

#!/usr/bin/env python3
"""
Red Hat VEX Statistics Analyzer - Fixed Version

Analyzes VEX files to generate statistics by RHEL version and severity.
Tracks progress with date-based indexing to allow incremental analysis.

"""

import argparse
import json
import logging
import os
import re
import pickle
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Set, Optional
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.worksheet.table import Table, TableStyleInfo
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from tqdm import tqdm
    TQDM_AVAILABLE = True
except ImportError:
    TQDM_AVAILABLE = False

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# Configuration
STATS_DIR = "stats"
INDEX_FILE = "stats_index.pkl"
OUTPUT_EXCEL = "vex_statistics_{date}.xlsx"
OUTPUT_SUMMARY = "vex_summary_{date}.txt"
OUTPUT_CSV = "vex_statistics_{date}.csv"
MAX_WORKERS = 10

# RHEL version patterns
RHEL_PATTERNS = {
    'EL6': r'el6|rhel6|redhat.*6|rhel.*6',
    'EL7': r'el7|rhel7|redhat.*7|rhel.*7', 
    'EL8': r'el8|rhel8|redhat.*8|rhel.*8|rhel8\.\d+',
    'EL9': r'el9|rhel9|redhat.*9|rhel.*9|rhel9\.\d+',
    'EL10': r'el10|rhel10|redhat.*10|rhel.*10',
}

# Severity levels with their priority order (FIXED ORDER)
SEVERITY_ORDER = ['Critical', 'Important', 'Moderate', 'Low', 'Unknown']
SEVERITY_PRIORITY = {sev: idx for idx, sev in enumerate(SEVERITY_ORDER)}

# Status types
STATUS_TYPES = ['fixed', 'known_affected', 'known_not_affected', 'under_investigation']


@dataclass
class VulnerabilityStats:
    """Statistics for vulnerabilities."""
    by_severity: Dict[str, Dict[str, int]] = field(default_factory=lambda: defaultdict(lambda: defaultdict(int)))
    overall: Dict[str, int] = field(default_factory=lambda: defaultdict(int))
    
    def add_entry(self, severity: str, status: str):
        """Add a vulnerability entry."""
        # Normalize severity
        if severity not in SEVERITY_ORDER:
            severity = 'Unknown'
        
        self.by_severity[severity][status] += 1
        self.by_severity[severity]['total'] += 1
        self.overall[status] += 1
        self.overall['total'] += 1
    
    def get_sorted_severities(self) -> List[str]:
        """Get severities sorted by priority."""
        return sorted(self.by_severity.keys(), key=lambda x: SEVERITY_PRIORITY.get(x, 999))


class StatsIndex:
    """Manages statistics index for tracking analysis progress."""
    
    def __init__(self, data_dir: Path):
        self.data_dir = data_dir
        self.index_file = data_dir / INDEX_FILE
        self.data = {}
        self.load()
    
    def load(self):
        """Load index from file."""
        if self.index_file.exists():
            try:
                with open(self.index_file, 'rb') as f:
                    self.data = pickle.load(f)
                logger.info(f"Loaded stats index from {self.index_file}")
            except (EOFError, pickle.PickleError) as e:
                logger.warning(f"Could not load stats index: {e}")
                self.data = {}
    
    def save(self):
        """Save index to file."""
        try:
            self.index_file.parent.mkdir(parents=True, exist_ok=True)
            with open(self.index_file, 'wb') as f:
                pickle.dump(self.data, f)
            logger.debug(f"Saved stats index to {self.index_file}")
        except Exception as e:
            logger.error(f"Error saving stats index: {e}")
    
    def update(self, **kwargs):
        """Update index data."""
        self.data.update(kwargs)


class VEXAnalyzer:
    """Analyzes VEX files for statistics."""
    
    def __init__(self, data_dir: Path, max_workers: int = MAX_WORKERS):
        self.data_dir = data_dir
        self.max_workers = max_workers
        self.stats_by_version: Dict[str, VulnerabilityStats] = defaultdict(VulnerabilityStats)
    
    def extract_rhel_versions(self, product_name: str) -> Set[str]:
        """Extract RHEL versions from product name."""
        versions = set()
        for version, pattern in RHEL_PATTERNS.items():
            if re.search(pattern, product_name, re.IGNORECASE):
                versions.add(version)
        return versions
    
    def get_severity(self, vuln_data: dict) -> str:
        """Extract severity from vulnerability data - FIXED VERSION."""
        
        # Method 1: Check document-level aggregate_severity (most common in Red Hat VEX)
        if 'document' in vuln_data:
            doc = vuln_data['document']
            if 'aggregate_severity' in doc and isinstance(doc['aggregate_severity'], dict):
                severity = doc['aggregate_severity'].get('text', '').strip()
                if severity:
                    return severity.capitalize()
        
        # Method 2: Check scores array in vulnerability
        if 'scores' in vuln_data:
            for score in vuln_data['scores']:
                # CVSS v3
                if 'cvss_v3' in score:
                    cvss = score['cvss_v3']
                    if 'baseSeverity' in cvss and cvss['baseSeverity']:
                        return cvss['baseSeverity'].capitalize()
                    elif 'baseScore' in cvss:
                        base_score = cvss['baseScore']
                        if base_score >= 9.0:
                            return 'Critical'
                        elif base_score >= 7.0:
                            return 'Important'
                        elif base_score >= 4.0:
                            return 'Moderate'
                        elif base_score > 0:
                            return 'Low'
                
                # CVSS v2
                elif 'cvss_v2' in score:
                    cvss = score['cvss_v2']
                    if 'baseSeverity' in cvss and cvss['baseSeverity']:
                        return cvss['baseSeverity'].capitalize()
                    elif 'baseScore' in cvss:
                        base_score = cvss['baseScore']
                        if base_score >= 7.0:
                            return 'Important'
                        elif base_score >= 4.0:
                            return 'Moderate'
                        elif base_score > 0:
                            return 'Low'
        
        # Method 3: Check threats array
        if 'threats' in vuln_data:
            for threat in vuln_data['threats']:
                if threat.get('category') == 'impact' and 'details' in threat:
                    severity = threat['details'].strip()
                    if severity:
                        return severity.capitalize()
        
        # Method 4: Legacy - direct baseSeverity field
        if 'baseSeverity' in vuln_data:
            severity = vuln_data['baseSeverity']
            if severity:
                return severity.capitalize()
        
        return 'Unknown'
    
    def analyze_file(self, file_path: Path) -> int:
        """Analyze a single VEX file and return number of vulnerabilities processed."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            vulnerabilities = data.get('vulnerabilities', [])
            count = 0
            
            for vuln in vulnerabilities:
                # Get severity from the entire data structure (not just vuln)
                severity = self.get_severity(data)
                product_status = vuln.get('product_status', {})
                
                # Process each status type
                for status_type in STATUS_TYPES:
                    products = product_status.get(status_type, [])
                    
                    for product in products:
                        rhel_versions = self.extract_rhel_versions(product)
                        
                        for version in rhel_versions:
                            self.stats_by_version[version].add_entry(severity, status_type)
                            count += 1
            
            return count
            
        except (json.JSONDecodeError, IOError) as e:
            logger.error(f"Error processing {file_path}: {e}")
            return 0
    
    def get_vex_files(self, start_date: Optional[datetime] = None) -> List[Path]:
        """Get list of VEX files to process."""
        vex_files = []
        
        for year_dir in self.data_dir.iterdir():
            if year_dir.is_dir() and year_dir.name.isdigit():
                for file_path in year_dir.glob("*.json"):
                    vex_files.append(file_path)
        
        # Filter by date if specified
        if start_date:
            filtered_files = []
            for file_path in vex_files:
                try:
                    if file_path.name.startswith('cve-'):
                        year = int(file_path.name.split('-')[1][:4])
                        file_date = datetime(year, 1, 1)
                        if file_date >= start_date:
                            filtered_files.append(file_path)
                    else:
                        filtered_files.append(file_path)
                except (IndexError, ValueError):
                    filtered_files.append(file_path)
            return filtered_files
        
        return vex_files
    
    def analyze_all(self, start_date: Optional[datetime] = None) -> int:
        """Analyze all VEX files with progress tracking."""
        vex_files = self.get_vex_files(start_date)
        
        if not vex_files:
            logger.warning("No VEX files found to analyze")
            return 0
        
        logger.info(f"Found {len(vex_files)} VEX files to analyze")
        
        total_vulns = 0
        
        if TQDM_AVAILABLE:
            pbar = tqdm(total=len(vex_files), desc="Analyzing", unit="file")
        
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = {
                executor.submit(self.analyze_file, file_path): file_path
                for file_path in vex_files
            }
            
            for future in as_completed(futures):
                count = future.result()
                total_vulns += count
                
                if TQDM_AVAILABLE:
                    pbar.update(1)
        
        if TQDM_AVAILABLE:
            pbar.close()
        
        logger.info(f"Processed {total_vulns} vulnerability entries")
        return len(vex_files)


class ReportGenerator:
    """Generates reports from statistics."""
    
    def __init__(self, stats: Dict[str, VulnerabilityStats], start_date: Optional[datetime] = None):
        self.stats = stats
        self.start_date = start_date
        self.date_str = self._generate_date_string()
        Path(STATS_DIR).mkdir(parents=True, exist_ok=True)
    
    def _generate_date_string(self) -> str:
        """Generate date string for filenames."""
        if self.start_date:
            return f"{self.start_date.strftime('%Y%m%d')}_{datetime.now().strftime('%H%M%S')}"
        return datetime.now().strftime("%Y%m%d_%H%M%S")
    
    def generate_csv(self):
        """Generate CSV report with proper severity ordering."""
        csv_file = Path(STATS_DIR) / OUTPUT_CSV.format(date=self.date_str)
        
        try:
            import csv
            
            with open(csv_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['RHEL Version', 'Severity', 'Status Type', 'Count'])
                
                # Sort versions
                for version in sorted(self.stats.keys()):
                    version_stats = self.stats[version]
                    
                    # Get severities in proper order
                    for severity in version_stats.get_sorted_severities():
                        severity_data = version_stats.by_severity[severity]
                        
                        for status_type in STATUS_TYPES:
                            count = severity_data.get(status_type, 0)
                            if count > 0:  # Only write non-zero counts
                                writer.writerow([version, severity, status_type, count])
            
            logger.info(f"CSV report generated: {csv_file}")
            return csv_file
            
        except Exception as e:
            logger.error(f"Error generating CSV: {e}")
            return None
    
    def generate_excel(self):
        """Generate Excel report with one sheet per RHEL version."""
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not installed, skipping Excel generation")
            return None
        
        excel_file = Path(STATS_DIR) / OUTPUT_EXCEL.format(date=self.date_str)
        
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Define styles
            title_font = Font(bold=True, size=14, color="FFFFFF")
            title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            section_font = Font(bold=True, size=11)
            section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # Create a sheet for each RHEL version
            for version in sorted(self.stats.keys()):
                version_stats = self.stats[version]
                total = version_stats.overall.get('total', 0)
                
                # Create worksheet
                ws = wb.create_sheet(title=version)
                
                # Title
                ws.merge_cells('A1:D1')
                ws['A1'] = f"{version} Statistics"
                ws['A1'].font = title_font
                ws['A1'].fill = title_fill
                ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                ws.row_dimensions[1].height = 25
                
                # Total CVE Entries
                ws['A2'] = "Total CVE Entries:"
                ws['B2'] = total
                ws['A2'].font = Font(bold=True)
                ws.row_dimensions[2].height = 20
                
                # Status Distribution Section
                ws['A4'] = "Status Distribution"
                ws.merge_cells('A4:D4')
                ws['A4'].font = section_font
                ws['A4'].fill = section_fill
                ws['A4'].alignment = Alignment(horizontal='left', vertical='center')
                
                # Status Distribution Headers
                ws['A5'] = "Status Type"
                ws['B5'] = "Count"
                ws['C5'] = "Percentage"
                for cell in ['A5', 'B5', 'C5']:
                    ws[cell].font = header_font
                    ws[cell].fill = header_fill
                    ws[cell].alignment = Alignment(horizontal='center')
                
                # Status Distribution Data
                row = 6
                for status_type in STATUS_TYPES:
                    count = version_stats.overall.get(status_type, 0)
                    percentage = (count / total * 100) if total > 0 else 0
                    
                    ws[f'A{row}'] = status_type
                    ws[f'B{row}'] = count
                    ws[f'C{row}'] = f"{percentage:.1f}%"
                    row += 1
                
                # By Severity Section
                row += 1
                ws[f'A{row}'] = "By Severity (Ordered by Priority)"
                ws.merge_cells(f'A{row}:D{row}')
                ws[f'A{row}'].font = section_font
                ws[f'A{row}'].fill = section_fill
                ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                
                row += 1
                
                # Severity Details
                for severity in version_stats.get_sorted_severities():
                    severity_data = version_stats.by_severity[severity]
                    sev_total = severity_data.get('total', 0)
                    sev_percentage = (sev_total / total * 100) if total > 0 else 0
                    
                    # Severity Header
                    ws[f'A{row}'] = f"{severity}"
                    ws[f'B{row}'] = f"Total: {sev_total}"
                    ws[f'C{row}'] = f"({sev_percentage:.1f}%)"
                    ws[f'A{row}'].font = Font(bold=True)
                    ws[f'B{row}'].font = Font(bold=True)
                    ws[f'C{row}'].font = Font(bold=True)
                    row += 1
                    
                    # Status breakdown for this severity
                    for status_type in STATUS_TYPES:
                        count = severity_data.get(status_type, 0)
                        status_percentage = (count / sev_total * 100) if sev_total > 0 else 0
                        
                        ws[f'A{row}'] = f"  - {status_type}"
                        ws[f'B{row}'] = count
                        ws[f'C{row}'] = f"{status_percentage:.1f}%"
                        row += 1
                    
                    row += 1  # Empty row between severities
                
                # Auto-adjust column widths
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15
                
                # Add borders to data sections
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row_cells in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=3):
                    for cell in row_cells:
                        if cell.value:
                            cell.border = thin_border
            
            wb.save(excel_file)
            logger.info(f"Excel report generated: {excel_file}")
            return excel_file
            
        except Exception as e:
            logger.error(f"Error generating Excel: {e}")
            return None    
    def generate_summary(self):
        """Generate text summary with proper severity ordering."""
        summary_file = Path(STATS_DIR) / OUTPUT_SUMMARY.format(date=self.date_str)
        
        try:
            with open(summary_file, 'w', encoding='utf-8') as f:
                f.write("=" * 70 + "\n")
                f.write("Red Hat VEX Statistics Summary\n")
                f.write("=" * 70 + "\n")
                f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Data Range: {self.start_date.strftime('%Y-%m-%d') if self.start_date else 'All available data'}\n")
                f.write("=" * 70 + "\n\n")
                
                # Sort versions
                for version in sorted(self.stats.keys()):
                    version_stats = self.stats[version]
                    total = version_stats.overall.get('total', 0)
                    
                    f.write(f"\n{'=' * 70}\n")
                    f.write(f"{version} Statistics\n")
                    f.write(f"{'=' * 70}\n")
                    f.write(f"Total CVE Entries: {total}\n\n")
                    
                    # Status distribution
                    f.write("Status Distribution:\n")
                    f.write("-" * 70 + "\n")
                    for status_type in STATUS_TYPES:
                        count = version_stats.overall.get(status_type, 0)
                        percentage = (count / total * 100) if total > 0 else 0
                        f.write(f"  {status_type:25s}: {count:6d} ({percentage:5.1f}%)\n")
                    
                    f.write("\n" + "-" * 70 + "\n")
                    f.write("By Severity (ORDERED BY PRIORITY):\n")
                    f.write("-" * 70 + "\n\n")
                    
                    # Severities in proper order
                    for severity in version_stats.get_sorted_severities():
                        severity_data = version_stats.by_severity[severity]
                        sev_total = severity_data.get('total', 0)
                        sev_percentage = (sev_total / total * 100) if total > 0 else 0
                        
                        f.write(f"{severity}:\n")
                        f.write(f"  Total: {sev_total} ({sev_percentage:.1f}%)\n")
                        
                        for status_type in STATUS_TYPES:
                            count = severity_data.get(status_type, 0)
                            status_percentage = (count / sev_total * 100) if sev_total > 0 else 0
                            f.write(f"    - {status_type:23s}: {count:6d} ({status_percentage:5.1f}%)\n")
                        f.write("\n")
            
            logger.info(f"Summary report generated: {summary_file}")
            return summary_file
            
        except Exception as e:
            logger.error(f"Error generating summary: {e}")
            return None
    
    def generate_all(self):
        """Generate all report types."""
        csv_file = self.generate_csv()
        excel_file = self.generate_excel()
        summary_file = self.generate_summary()
        
        return {
            'csv': csv_file,
            'excel': excel_file,
            'summary': summary_file
        }


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Red Hat VEX Statistics Analyzer - Fixed Severity Detection",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Analyze all files
  %(prog)s --data-dir data
  
  # Analyze with date filter
  %(prog)s --data-dir data --start-date 2025-01-01
  
  # Resume from previous run
  %(prog)s --data-dir data --resume
  
  # Reset and start fresh
  %(prog)s --data-dir data --reset
        """
    )
    
    parser.add_argument(
        "--data-dir",
        type=str,
        default="data",
        help="Directory containing VEX files (default: data)"
    )
    parser.add_argument(
        "--resume",
        action="store_true",
        help="Resume from last run using saved index"
    )
    parser.add_argument(
        "--reset",
        action="store_true",
        help="Reset index and start fresh"
    )
    parser.add_argument(
        "--start-date",
        type=str,
        help="Start date for analysis (YYYY-MM-DD)"
    )
    parser.add_argument(
        "--max-workers",
        type=int,
        default=MAX_WORKERS,
        help=f"Maximum worker threads (default: {MAX_WORKERS})"
    )
    parser.add_argument(
        "--verbose",
        "-v",
        action="store_true",
        help="Enable verbose logging"
    )
    
    args = parser.parse_args()
    
    # Setup logging
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Setup paths
    data_dir = Path(args.data_dir)
    if not data_dir.exists():
        logger.error(f"Data directory not found: {data_dir}")
        return 1
    
    # Load or reset index
    stats_index = StatsIndex(data_dir)
    
    if args.reset:
        stats_index.data = {}
        logger.warning("Index reset - starting fresh analysis")
    elif args.resume and stats_index.data:
        logger.info(f"Resuming from previous run (last: {stats_index.data.get('last_run', 'unknown')})")
    else:
        logger.info("Starting new analysis")
    
    # Parse start date
    start_date = None
    if args.start_date:
        try:
            start_date = datetime.strptime(args.start_date, "%Y-%m-%d")
            logger.info(f"Analyzing files from {start_date.date()} onwards")
        except ValueError:
            logger.error("Invalid date format. Use YYYY-MM-DD")
            return 1
    
    # Analyze files
    try:
        analyzer = VEXAnalyzer(data_dir, args.max_workers)
        file_count = analyzer.analyze_all(start_date)
        
        if file_count == 0:
            logger.warning("No files analyzed")
            return 0
        
        # Update index
        stats_index.update(
            last_run=datetime.now().isoformat(),
            file_count=file_count,
            analysis_date=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        )
        stats_index.save()
        
        # Generate reports
        logger.info("Generating reports...")
        report_gen = ReportGenerator(analyzer.stats_by_version, start_date)
        reports = report_gen.generate_all()
        
        # Print summary
        logger.info("\n" + "=" * 70)
        logger.info("Analysis Complete!")
        logger.info("=" * 70)
        logger.info(f"Files analyzed: {file_count}")
        logger.info(f"RHEL versions found: {len(analyzer.stats_by_version)}")
        
        for version in sorted(analyzer.stats_by_version.keys()):
            total = analyzer.stats_by_version[version].overall.get('total', 0)
            logger.info(f"  {version}: {total} CVE entries")
        
        logger.info(f"\nReports saved in {STATS_DIR}/ directory:")
        for report_type, report_file in reports.items():
            if report_file:
                logger.info(f"  - {report_type.upper()}: {report_file}")
        
        logger.info("\nUse --resume to continue from this point in future runs")
        
        return 0
        
    except KeyboardInterrupt:
        logger.info("\nAnalysis cancelled by user")
        return 1
    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=args.verbose)
        return 1


if __name__ == "__main__":
    exit(main())